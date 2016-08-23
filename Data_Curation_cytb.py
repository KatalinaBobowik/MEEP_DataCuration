from Bio import Entrez
from Bio import SeqIO
from openpyxl import Workbook
import requests, os
import shutil

# to get the relative path
FASTA = os.path.join(os.path.abspath("."), u"fasta")

# {'name': 'search_string_1'}

def header():
    print("-----------------------")
    print("  DATA CURATION SCRIPT")
    print("-----------------------")


def get_user_input():

    if not os.path.exists(FASTA):
            os.makedirs(FASTA)
    else:
        fasta = os.listdir(FASTA)
        for each in fasta:
            os.remove(FASTA + "/" + each)

    keywords = {}

    global organism_names

    organism_names = ["Alopex lagopus", "Bos primigenius", "Bison priscus",
                         "Balaena mysticetus","Ursus spelaeus", "Crocuta crocuta spelaea",
                         "Panthera leo spelaea","Zea mays", "Pachyornis mappini",
                         "Ovibos moschatus","Ctenomys sociabilis", "Mammuthus primigenius",
                         "Vulpes vulpes","Hatteria punctata", "Microtus arvalis", "Phytophthora infestans",
                         ]
    # test case
    #organism_names = ["Bos primigenius"]


    for name in organism_names:

        keywords[name] = '"{}"[Organism] AND ((cytb[Gene Name]) OR (cytochrome b[Gene Name])) AND 100:1000[SLEN] AND biomol_genomic[PROP]'.format(name)

    email = raw_input("Enter an email address: ")
    database = raw_input("Enter a database name (eg: nucleotide): ")

    if email and database:
        Entrez.email = email
        ncbi_input(database, keywords)

    else:

        print ""
        print "Sorry, you must enter a value "
        print "----------"
        print ""

        get_user_input()


def ncbi_input(database, keywords):

    print "\nSearching in %s..." %(database)

    handle_list = {}

    for name in keywords.keys():

        search_string = keywords[name]
        print search_string

        # retmax == maximum number of return results. In this instance, no retmax value specified.
        raw_xml_return = Entrez.esearch(db=database, term=search_string, retmax = 1000)

        transform_xml = Entrez.read(raw_xml_return)
        print "Acquired {} of {}".format(len(transform_xml["IdList"]), transform_xml["Count"])

        handle_list[name] = transform_xml

    if len(handle_list) != 0:
        download_fasta(handle_list)

    else:
        print "Nothing was found using your input keywords"


def n_split(iterable, n, fillvalue=None):

    num_extra = len(iterable) % n
    zipped = zip(*[iter(iterable)] * n)
    return zipped if not num_extra else zipped + [iterable[-num_extra:], ]

# download the data for each GI number
def download_fasta(handle_list):

    print "\nDownloading fasta data..."
    print ""
    result = {}

    for each in handle_list.keys():

        genbankidlist = handle_list[each]['IdList']

        count = 1
        for gen_id in n_split(genbankidlist, 200):

            handle = Entrez.efetch(db='nucleotide', id=list(gen_id), rettype='fasta')

            # to get the author we can use the following:
            #handle = Entrez.efetch(db="nucleotide",id="635545179",rettype="gb", retmode="xml")

            print "Our URL is {}".format(handle.url)

            download_file(handle.url, each + "_{}.fasta".format(count))
            count+=1

            handle.close()

    merge_files()

    create_excel_output()


def merge_files():

    #path to the fasta files
    path = os.getcwd() + '/fasta/'

    aggregated_files = {}

    #prepare a dictionary of names as keys with corresponding empty lists
    for organism in organism_names:
        aggregated_files[organism] = []

    #match and append filename to organism name
    for name in aggregated_files:
        for each_file in os.listdir(path):
            if name in each_file:
                aggregated_files[name].append(each_file)

    for name in aggregated_files:

        # destination = open(path + name + '.fasta', 'wb')
        with open(path + name + '.fasta', 'wb') as destination:

            for filename in aggregated_files[name]:
                shutil.copyfileobj(open(path + filename, 'rb'), destination)
                os.remove(path + filename)


def download_file(url, filename):

    # NOTE the stream=True parameter
    print "Writing {}".format(filename)
    r = requests.get(url, stream=True)

    with open(FASTA + "/" + filename, 'wb') as f:
        for chunk in r.iter_content(chunk_size=1024):
            if chunk: # filter out keep-alive new chunks
                f.write(chunk)
                f.flush()
                #f.flush() commented by recommendation from J.F.Sebastian

    print "Finished writing"
    print "----------------"


def run():
    header()
    get_user_input()


def create_excel_output():

    wb = Workbook()

    files = os.listdir(FASTA)


    for name in files:
        if ".fasta" in name:

            sheet_name = name.split(".")[0]
            ws = wb.active
            ws = wb.create_sheet(title=sheet_name)
            ws.cell(column=1, row=1, value="Genbank sequence name")
            ws.cell(column=2, row=1, value="Name & Region")
            ws.cell(column=3, row=1, value="Accession Number")
            ws.cell(column=4, row=1, value="Genbank ID")
            ws.cell(column=5, row=1, value="Radiocarbon Date")
            ws.cell(column=6, row=1, value="Title")
            ws.cell(column=7, row=1, value="FASTA Link")

            count = 2
            #open_fasta = SeqIO.parse(open(FASTA + "/" + name, "rU"), "fasta")
            with open(FASTA + "/" + name, "rU") as open_fasta:

                parsed_fasta = SeqIO.parse(open_fasta, "fasta")

                for sequence in parsed_fasta:

                    seq_name = sequence.name
                    latin_name_region = sequence.description.split("|")[4]
                    accession = sequence.id.split("|")[3]
                    gb_id = sequence.id.split("|")[1]
                    link_fasta = FASTA + "/" + name

                    print latin_name_region
                    print accession
                    print gb_id

                    handle = Entrez.efetch(db="nucleotide",id=gb_id, rettype="gb", retmode="text")
                    record = list(SeqIO.parse(handle, "genbank"))[0]

                    #see if excel can handle two records
                    paper_title = record.annotations["references"][0].title

                    ws.cell(column=1, row=count, value=seq_name)
                    ws.cell(column=2, row=count, value=latin_name_region)
                    ws.cell(column=3, row=count, value=accession)
                    ws.cell(column=4, row=count, value=gb_id)
                    ws.cell(column=6, row=count, value=paper_title)
                    ws.cell(column=7, row=count, value=link_fasta)
                    count= count + 1


    ws = wb.get_sheet_by_name("Sheet")
    wb.remove_sheet(ws)
    wb.save(filename = "Data_Curation.xlsx")

run()

#ValueError: More than one record found in handle
#error at Sphenodon punctatus isolate FT2522 Tawhiti Rahi mitochondrial control region, partial sequence AF442421.1, id="1699339"


#pull d-loop from search
# total of 300
